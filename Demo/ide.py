# coding:utf-8
import xlrd
import psutil
import xlutils
import xlwt
import openpyxl
import os
import pandas
import sys
import datetime

# c=datetime.datetime.now()
# print(c)
# d=datetime.datetime.date(c)
# print(d)

path=os.getcwd()
print(path)
t=psutil.disk_usage(path).total/1024/1024/1024
u=psutil.disk_usage(path).used/1024/1024/1024
f=psutil.disk_usage(path).free/1024/1024/1024
p=psutil.disk_usage(path).percent
print(f)
print(u)
print(t)
print(p)

# base_path = '2021-09-29 10_29_26-cashApiInterFace数据.xlsx'
# ob = openpyxl.load_workbook(filename=base_path)
# sn = ob.sheetnames
# sh = ob[sn[0]]
# li = []
# for i in sh[47]:
#     li.append(i.value)
# print(li)
# print(li[1])
# print(type(li[1]))
# print(li[1].strip("/"))
# print(li)
# path=os.getcwd()
# print(path)
# disk=psutil.disk_usage(path)
# f=disk.free/1024/1024/1024
# print(f)
# print(disk)



# class HandleExcel:
#     def load_excel(self):
#         '''
#         加载excel
#         '''
#         open_excel = openpyxl.load_workbook(base_path)#拿到excel的所有内容
#         return open_excel
#     def get_sheet_data(self,index=None):
#         '''
#         加载所有sheet的内容
#         '''
#         sheet_name = self.load_excel().sheetnames#拿到sheetnames的所有内容
#         if index == None:
#             index = 0
#         data = self.load_excel()[sheet_name[index]]
#         return data
#     def get_cell_value(self,row,cols):
#         '''
#         获取某一个单元格内容
#         '''
#         data = self.get_sheet_data().cell(row=row,column=cols)
#         return data
#     def get_rows(self):
#         row = self.get_sheet_data().max_row
#         return row
#     def get_rows_value(self,row):
#         '''
#         获取某一行的内容
#         '''
#         row_list = []
#         for i in self.get_sheet_data()[row]:
#             row_list.append(i.value)
#         return row_list
#
# if __name__ == '__main__':
#     handle = HandleExcel()
#     print(handle.get_rows_value(2))
